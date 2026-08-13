#!/usr/bin/env python3
"""
多语言Excel翻译工具 v2.0
=========================
功能：读取Excel文档，以中文（B列）为源语言，自动补全11种目标语言中缺失的翻译。
     已有翻译的单元格不会被覆盖。使用MyMemory翻译API（免费，无需API密钥）。

目标语言：英语、繁体中文、日语、韩语、德语、西班牙语、俄语、法语、葡萄牙语、意大利语、泰语

用法：
    translate_excel.exe <输入文件.xlsx> [输出文件.xlsx]
    translate_excel.exe 多语言_已翻译.xlsx
    translate_excel.exe input.xlsx output.xlsx
"""

import sys
import os
import re
import time
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import requests

# ---- 配置 ----
SOURCE_COL = "B"

# 目标语言：列 -> (MyMemory语言对后半部分, 语言名)
LANG_MAP = {
    "C": ("en",     "英语"),
    "D": ("zh-TW",  "繁体中文"),
    "E": ("ja",     "日语"),
    "F": ("ko",     "韩语"),
    "G": ("de",     "德语"),
    "H": ("es",     "西班牙语"),
    "I": ("ru",     "俄语"),
    "J": ("fr",     "法语"),
    "K": ("pt",     "葡萄牙语"),
    "L": ("it",     "意大利语"),
    "M": ("th",     "泰语"),
}

MYMEMORY_URL = "https://api.mymemory.translated.net/get"
TRANSLATE_DELAY = 0.4
MAX_RETRIES = 3


def parse_xlsx(filepath):
    """使用XML解析xlsx，兼容损坏的table引用"""
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    with zipfile.ZipFile(filepath, 'r') as z:
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            tree = ET.parse(z.open('xl/sharedStrings.xml'))
            for si in tree.findall('.//s:si', ns):
                text = ''.join(t.text or '' for t in si.findall('.//s:t', ns))
                shared_strings.append(text)

        wb_tree = ET.parse(z.open('xl/workbook.xml'))
        sheets = [sh.get('name') for sh in wb_tree.findall('.//s:sheet', ns)]

        all_data = {}
        for idx, sheet_name in enumerate(sheets):
            sheet_file = f'xl/worksheets/sheet{idx + 1}.xml'
            if sheet_file not in z.namelist():
                continue
            tree = ET.parse(z.open(sheet_file))
            rows = tree.findall('.//s:row', ns)
            sheet_data = []
            for row in rows:
                row_data = {}
                for c in row.findall('s:c', ns):
                    ref = c.get('r')
                    col_match = re.match(r'([A-Z]+)', ref)
                    if not col_match:
                        continue
                    col_letter = col_match.group(1)
                    t = c.get('t')
                    v = c.find('s:v', ns)
                    val = v.text if v is not None else ''
                    if t == 's' and val.isdigit():
                        idx_s = int(val)
                        val = shared_strings[idx_s] if idx_s < len(shared_strings) else val
                    row_data[col_letter] = val
                sheet_data.append(row_data)
            all_data[sheet_name] = sheet_data

        return all_data


def translate_text(text, target_lang):
    """调用MyMemory API翻译文本，提取最佳机器翻译结果"""
    langpair = f"zh-CN|{target_lang}"
    params = {'q': text, 'langpair': langpair, 'mt': '1'}

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.get(MYMEMORY_URL, params=params, timeout=15)
            if resp.status_code != 200:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2)
                    continue
                return ''

            data = resp.json()

            # 优先取神经机器翻译(MT)结果
            for m in data.get('matches', []):
                if m.get('model') == 'neural' or m.get('created-by') == 'MT!':
                    return m.get('translation', '')

            # 备用：取responseData中的翻译
            translated = data.get('responseData', {}).get('translatedText', '')
            return translated or ''

        except Exception:
            if attempt < MAX_RETRIES - 1:
                time.sleep(2)
            else:
                return ''

    return ''


def print_summary(all_data):
    """打印翻译缺失概览"""
    print("\n" + "=" * 60)
    print("  Excel 数据概览")
    print("=" * 60)
    total_missing = 0
    total_cells = 0

    for sheet_name, sheet_data in all_data.items():
        cn_count = sum(1 for r in sheet_data if r.get(SOURCE_COL, '').strip())
        if cn_count == 0:
            continue

        missing_detail = {}
        for col_letter, (_, lang_name) in LANG_MAP.items():
            m = sum(1 for r in sheet_data
                    if r.get(SOURCE_COL, '').strip() and not r.get(col_letter, '').strip())
            if m > 0:
                missing_detail[lang_name] = m
        total_cells += cn_count * len(LANG_MAP)
        total_missing += sum(missing_detail.values())

        print(f"\n  [{sheet_name}]: {cn_count} 条中文")
        if missing_detail:
            for lang, cnt in missing_detail.items():
                print(f"    - {lang}: {cnt} 条缺失")
        else:
            print(f"    [OK] 翻译完整")

    print(f"\n  总计缺失: {total_missing}/{total_cells} 个单元格")
    print("=" * 60)
    return total_missing


def write_xlsx(all_data, output_path):
    """用openpyxl写入xlsx（A-M列）"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        ('A', 'ID'), ('B', '中文'), ('C', '英语'), ('D', '繁体'),
        ('E', '日语'), ('F', '韩语'), ('G', '德语'), ('H', '西班牙语'),
        ('I', '俄语'), ('J', '法语'), ('K', '葡萄牙语'), ('L', '意大利语'),
        ('M', '泰语')
    ]

    for sheet_name, sheet_data in all_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, (_, header_name) in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header_name)

        for row_offset, row_data in enumerate(sheet_data):
            excel_row = row_offset + 2
            for col_idx, (col_letter, _) in enumerate(headers, 1):
                val = row_data.get(col_letter, '')
                if val:
                    ws.cell(row=excel_row, column=col_idx, value=val)

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        input("\n按回车键退出...")
        sys.exit(0)

    input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        stem = Path(input_file).stem
        suffix = Path(input_file).suffix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = str(Path(input_file).parent / f"{stem}_已翻译_{timestamp}{suffix}")

    if not os.path.exists(input_file):
        print(f"[ERROR] 文件不存在: {input_file}")
        input("\n按回车键退出...")
        sys.exit(1)

    print(f"\n  Input : {input_file}")
    print(f"  Output: {output_file}")

    # 解析
    print("\n[1/4] 解析 Excel...")
    all_data = parse_xlsx(input_file)

    # 分析
    print("[2/4] 分析缺失翻译...")
    missing_count = print_summary(all_data)
    if missing_count == 0:
        print("\n[OK] 无需翻译！")
        input("\n按回车键退出...")
        return

    # 收集任务
    tasks = []
    for sheet_name, sheet_data in all_data.items():
        for row_idx, row_data in enumerate(sheet_data):
            src = row_data.get(SOURCE_COL, '').strip()
            if not src:
                continue
            for col_letter, (lang_code, lang_name) in LANG_MAP.items():
                if not row_data.get(col_letter, '').strip():
                    tasks.append((sheet_name, row_idx, col_letter, src, lang_code, lang_name))

    total = len(tasks)
    print(f"\n[3/4] 开始翻译 ({total} 条)...\n")

    # 按语言分组翻译
    by_lang = defaultdict(list)
    for t in tasks:
        by_lang[t[4]].append(t)

    done = 0
    for lang_code, group in sorted(by_lang.items()):
        lang_name = group[0][5]
        cnt = len(group)
        print(f"  [{lang_name}] {cnt} 条")

        for i, (sheet_name, row_idx, col_letter, src, lc, ln) in enumerate(group):
            result = translate_text(src, lc)
            all_data[sheet_name][row_idx][col_letter] = result
            done += 1

            if (i + 1) % 5 == 0 or (i + 1) == cnt:
                pct = done * 100 // total
                print(f"    进度: {done}/{total} ({pct}%)")

            time.sleep(TRANSLATE_DELAY)

    # 保存
    print(f"\n[4/4] 保存结果...")
    write_xlsx(all_data, output_file)

    print(f"\n{'=' * 60}")
    print(f"  [DONE] 翻译完成！共翻译 {done} 个单元格")
    print(f"  输出: {os.path.abspath(output_file)}")
    print(f"{'=' * 60}")
    input("\n按回车键退出...")


if __name__ == '__main__':
    main()
